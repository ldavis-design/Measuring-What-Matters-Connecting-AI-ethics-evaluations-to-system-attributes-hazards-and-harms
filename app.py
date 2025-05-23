import os
import pandas as pd
import dash
from dash import dcc, html, Input, Output
import plotly.express as px
from openpyxl import load_workbook

# Load the dataset
dropped_df = pd.read_excel("RAI_Measures_Dataset.xlsx")
dropped_df.columns = dropped_df.iloc[0]  # Use second row as column names
dropped_df = dropped_df.iloc[1:]  # Drop first two rows

wb = load_workbook("RAI_Measures_Dataset.xlsx", data_only=True)
ws = wb.active

link_map = {}
for row in ws.iter_rows(min_row=3):  # adjust if headers differ
    title_cell = row[dropped_df.columns.get_loc("Title")]
    if title_cell.hyperlink:
        link_map[title_cell.value] = title_cell.hyperlink.target

# Map the hyperlinks into a new column
dropped_df["Access Link"] = dropped_df["Title"].map(link_map)

# Subset and prepare data
subset_df_process = dropped_df[
    ['Principle', 'Component of the ML System', 'Measure', 'Measurement Process', 'Title',
     'Type of Assessment', 'Application Area', 'Year', 'Primary Harm', 'Secondary Harm',
     'Attribute', 'Hazard', 'Access Link']
]

# Group and clean data
grouped_df_process = subset_df_process.groupby(
    ['Principle', 'Component of the ML System', 'Measurement Process', 'Title', 'Primary Harm',
     'Secondary Harm', 'Attribute', 'Hazard', 'Access Link', 'Type of Assessment', 'Application Area', 'Year']
)['Measure'].apply(list).reset_index()

grouped_df_process['Principle'] = grouped_df_process['Principle'].astype(str).str.split(', ')
grouped_df_process['Component of the ML System'] = grouped_df_process['Component of the ML System'].astype(str).str.split(', ')
grouped_df_process['Primary Harm'] = grouped_df_process['Primary Harm'].astype(str).str.split(', ')

grouped_df_process = grouped_df_process.explode('Principle')
grouped_df_process = grouped_df_process.explode('Component of the ML System')
grouped_df_process = grouped_df_process.explode('Primary Harm')

grouped_df_process['Principle'] = grouped_df_process['Principle'].str.strip()
grouped_df_process['Component of the ML System'] = grouped_df_process['Component of the ML System'].str.strip()
grouped_df_process['Primary Harm'] = grouped_df_process['Primary Harm'].str.strip()

grouped_df_process['Measure'] = grouped_df_process['Measure'].apply(lambda x: x if isinstance(x, list) else [x])
grouped_df_process = grouped_df_process.explode('Measure')
grouped_df_process['Measure'] = grouped_df_process['Measure'].str.strip()

# Remove rows with nulls in sunburst path
grouped_df_process = grouped_df_process.dropna(subset=[
    'Principle', 'Component of the ML System', 'Primary Harm', 'Measure'
])

# Define custom color palette
custom_palette = [
    "#9c0040", "#ff7e3c","#ff3d54", "#ffc68e", "#e9e807",
    "#87ed2d", "#66c2a5", "#3288bd", "#5e4fa2", "#0a2e58",
    "#adadad"
]

# Create sunburst chart
fig = px.sunburst(
    grouped_df_process,
    path=["Principle", "Component of the ML System", "Primary Harm", "Measure"],
    values=None,
    title=" ",
    color="Principle",
    color_discrete_sequence=custom_palette,
    custom_data=["Principle"]
)

fig.update_traces(
    hovertemplate="<b>%{label}</b><br>Parent: %{parent}<br>Principle: %{customdata[0]}<br><extra></extra>"
)

fig.update_layout(
    margin=dict(l=20, r=20, t=30, b=20),
    height=700,
    width=1500,
    legend=dict(
        orientation="h",
        yanchor="bottom",
        y=1.02,
        xanchor="right",
        x=1,
        tracegroupgap=10,
        font=dict(family="Source Sans Pro, Arial, sans-serif")
    ),
    uniformtext_minsize=8,
    font=dict(family="Source Sans Pro, Arial, sans-serif")
)

# Initialize Dash app
app = dash.Dash(__name__)

app.layout = html.Div(
    style={"font-family": "Source Sans Pro, Arial, sans-serif", "margin": "20px"},
    children=[
        html.H1(
            "Responsible AI Measures Dataset",
            style={"text-align": "center", "color": "#151417"}
        ),
        html.Div(
            id='click-output',
            style={
                "margin-bottom": "20px",
                "font-size": "16px",
                "color": "#151417",
                "text-align": "left"
            }
        ),
        dcc.Graph(
            id='sunburst-chart',
            figure=fig,
            config={'displayModeBar': False}
        )
    ]
)

@app.callback(
    Output('click-output', 'children'),
    Input('sunburst-chart', 'clickData')
)
def display_click_data(clickData):
    if clickData is None:
        return html.Div([
            html.B("Instructions for Use:"),
            html.Ul([
                html.Li("Please select a principle, followed by the component of the ML system that you are interested in exploring, and the primary harm."),
                html.Li("Hovering will display three pieces of metadata: the tier you are currently at in the visual, the parent (e.g., the tier prior), and the principle you are currently hovering above."),
                html.Li("Click on a measure to see the corresponding measurement process."),
                html.Li("To learn more about the measure, its formulaic variables (if quantitative), secondary harm (if present), hazard, attribute, and relative context of use, please click the link on each measurement process to access the authors’ publication.")
            ]),
            html.I("Some measurement processes will include paper-specific references, terms, or formulas that may require further context to understand. Please use the paper title and lead author name(s) to further investigate the measure(s).")
        ])

    point_data = clickData.get('points', [{}])[0]
    clicked_label = point_data.get('label')

    if not clicked_label:
        return html.Span("Please click on an RAI Measure in the last tier.")

    if clicked_label in grouped_df_process['Measure'].values:
        row = grouped_df_process.loc[grouped_df_process['Measure'] == clicked_label].iloc[0]

        return html.Div([
            html.B(f"{clicked_label}: "), html.Span(row['Measurement Process']),
            html.Br(), html.Br(),
            html.B("Paper Title: "),
            html.A(row['Title'], href=row['Access Link'], target="_blank", style={
                "text-decoration": "underline", "color": "#1a0dab"
            }),
            html.Br(),
            html.B("Publication Year: "), html.Span(row['Year']),
            html.Br(), html.Br(),
            html.B("Hazard: "), html.Span(row['Hazard']),
            html.Br(),
            html.B("Secondary Harm: "), html.Span(row['Secondary Harm']),
            html.Br(),
            html.B("Attribute: "), html.Span(row['Attribute']),
            html.Br(), html.Br(),
            html.B("Type of Assessment: "), html.Span(row['Type of Assessment']),
            html.Br(),
            html.B("Application Area: "), html.Span(row['Application Area']),
            html.Br(), html.Br(),
            html.I("As stated in the initial informational screen, some measurement processes will include paper-specific references, terms, or formulas that may require further context to understand. Please use the paper title and lead author name(s) to further investigate the measure(s).")
        ])
    else:
        return html.Span(
            "Please click on an RAI Measure in the last tier.",
            style={"font-family": "Source Sans Pro, Arial, sans-serif"}
        )

# Run app
server = app.server

if __name__ == '__main__':
    port = int(os.environ.get("PORT", 8051))  # Default to port 8051
    app.run(host="0.0.0.0", port=port, debug=True)
